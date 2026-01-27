import sys
import openpyxl
import shutil
from datetime import datetime, timedelta
from openpyxl.utils import range_boundaries
import os
import platform

# 定义需要复制的数据范围
SELL_RANGES = [(6, 16), (18, 25), (27, 35), (37, 38), (40, 48), (50, 52), (54, 55), (57, 59)]
ROW_RANGES = [(65, 75), (77, 84), (86, 94), (96, 97), (99, 107), (109, 111), (113, 114), (116, 118)]

def format_chinese_date(dt):
    """跨平台兼容的中文日期格式（1月8日）"""
    year = dt.year
    month = dt.month  # 自动去掉前导0
    day = dt.day
    return f"{month}月{day}日",f"{month}.{day}",f"{year}{month}{day}"

def copy_file_1(template_file1):
    """
    复制模板文件，在文件名前添加昨天日期
    返回新文件路径，失败返回None
    """
    try:
        template_path = os.path.abspath(template_file1)
		
        if not os.path.exists(template_path):
            print(f"❌ 错误: 模板文件不存在: {template_path}")
            return None
            
        template_dir = os.path.dirname(template_path)
        
        # 获取昨天日期
        yesterday = datetime.now() - timedelta(days=1)
        data1,data2,data3 = format_chinese_date(yesterday)
        
        # 生成新文件名（日期前缀）
        new_filename = f"{data1}内衣销售数据.xlsx"
        src_filename = f"双澄内衣日报{data2}.xlsx"
        new_path = f"D:/workspace/Nancy/to/{new_filename}"
        src_file = f"D:/workspace/Nancy/from/{src_filename}"
        # 如果目标文件已存在，先删除
        if os.path.exists(new_path):
            os.remove(new_path)
            print(f"🗑️  已删除已存在文件: {new_filename}")
			
        if os.path.exists(src_file):
            print(f"🗑️  文件不存在: {src_filename}")
			
        if not os.path.exists(template_dir):
            print(f"❌ 错误: 模板文件不存在: {template_path}")
            return None
        # 复制文件
        shutil.copy2(template_path, new_path)
        print(f"✅ 成功创建文件: {new_filename}")
        print(f"📅 日期: {data1}")
        
        return new_path,src_file
        
    except Exception as e:
        print(f"❌ 复制模板失败: {str(e)}")
        return None

def copy_file_2(template_file2):
    try:
        template_path = os.path.abspath(template_file2)
		
        if not os.path.exists(template_path):
            print(f"❌ 错误: 模板文件不存在: {template_path}")
            return None
            
        template_dir = os.path.dirname(template_path)
        template_name = os.path.basename(template_path)
        
        # 获取昨天日期
        yesterday = datetime.now() - timedelta(days=1)
        data1,data2,data3 = format_chinese_date(yesterday)
        
        # 生成新文件名（日期前缀）
        new_filename = f"内衣GSV{data3}.xlsx"
        src_filename = f"GSV双澄内衣日报{data2}.xlsx"
        new_path = f"D:/workspace/Nancy/to/{new_filename}"
        src_file = f"D:/workspace/Nancy/from/{src_filename}"
        # 如果目标文件已存在，先删除
        if os.path.exists(new_path):
            os.remove(new_path)
            print(f"🗑️  已删除已存在文件: {new_filename}")
			
        if os.path.exists(src_file):
            print(f"🗑️  文件不存在: {src_filename}")
			
        if not os.path.exists(template_dir):
            print(f"❌ 错误: 模板文件不存在: {template_path}")
            return None
        # 复制文件
        shutil.copy2(template_path, new_path)
        print(f"✅ 成功创建文件: {new_filename}")
        print(f"📅 日期: {data1}")
        
        return new_path,src_file
        
    except Exception as e:
        print(f"❌ 复制模板失败: {str(e)}")
        return None

def copy_ranges(src_column,dst_column,src_path, dst_path):
    """
    从源文件复制L列(第12列)数据到目标文件的C列(第3列)
    按SELL_RANGES定义的行范围依次复制
    """
    src_path = os.path.abspath(src_path)
    dst_path = os.path.abspath(dst_path)

    if not os.path.isfile(src_path):
        raise FileNotFoundError(f'源文件不存在：{src_path}')

    # 确保输出目录存在
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)

    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src.active

    # 加载或创建工作簿
    if os.path.isfile(dst_path):
        wb_dst = openpyxl.load_workbook(dst_path)
    else:
        wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active

    # 复制数据：源L列(12) → 目标C列(3)
    dst_row = 2
    count = 0
    for start, end in SELL_RANGES:
        for r in range(start, end + 1):
            value = ws_src.cell(row=r, column=src_column).value
            ws_dst.cell(row=dst_row, column=dst_column, value=value)
            dst_row += 1
            count += 1

    wb_dst.save(dst_path)
    wb_src.close()
    wb_dst.close()
    print(f'✅ 已复制 {count} 条销售数据 → {dst_column}列')

def pick_one_per_row(ws, row):
    """
    在J(10)/K(11)/L(12)三列中查找唯一非空值
    优先级: J > K > L
    """
    cols = [10, 11, 12]  # J, K, L
    for col in cols:
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip() != '':
            return val
    return None

def copy_j_ranges(src_path, dst_path):
    """
    1. 从源文件J/K/L列提取有效值，写入目标文件D列(第4列)
    2. 在B列(第2列)填充昨天的日期(YYYYMMDD格式)
    """
    src_path = os.path.abspath(src_path)
    if not os.path.isfile(src_path):
        raise FileNotFoundError(f'源文件不存在：{src_path}')

    # 收集数据
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src.active
    
    values = []
    for start, end in ROW_RANGES:
        for r in range(start, end + 1):
            val = pick_one_per_row(ws_src, r)
            values.append(val)
    
    wb_src.close()

    # 写入目标文件
    if os.path.isfile(dst_path):
        wb_dst = openpyxl.load_workbook(dst_path)
    else:
        wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active

    # 写入D列（从第2行开始）
    for idx, val in enumerate(values, start=2):
        ws_dst.cell(row=idx, column=4, value=val)

    # 在B列填充昨天日期（与数据行数一致）
    yesterday_num = int((datetime.now() - timedelta(days=1)).strftime('%Y%m%d'))
    for r in range(2, 2 + len(values)):
        ws_dst.cell(row=r, column=2, value=yesterday_num)
    
    print(f"📅 填充日期: {yesterday_num} (共{len(values)}行)")

    wb_dst.save(dst_path)
    wb_dst.close()
    print(f'✅ 已写入 {len(values)} 行其他数据 → D列')

def adjust_file2_data(dst_path):
    # 1. 计算昨天日期
    yesterday = datetime.now() - timedelta(days=1)
    # 2. 加载 Excel 文件
    wb = openpyxl.load_workbook(dst_path)
    ws = wb.active  # 或使用 ws = wb["Sheet1"] 指定工作表名

    # 3. 在 A 列第 2~48 行填充昨天日期
    for row in range(2, 49):  # range(2, 49) 表示 2 到 48（含）
        cell = ws.cell(row=row, column=1)  # column=1 即 A 列
        cell.value = yesterday
        cell.number_format = 'YYYY/M/D'  # 设置显示格式为 2026/1/23（不带前导零）

    # 4. 保存
    wb.save(dst_path)
    print(f"✅ 已完成：A2:A48 已填充 {yesterday.strftime('%Y/%m/%d')}（昨天）")

def main():
    print("Excel 数据处理工具")
    print("=" * 50)
    
    # 步骤1：复制模板（带昨天日期）
    template_file1 = "D:/workspace/Nancy/example/内衣销售数据.xlsx"
    result_path1, source_file1= copy_file_1(template_file1)
	
    template_file2 = "D:/workspace/Nancy/example/内衣GSV.xlsx"
    result_path2, source_file2= copy_file_2(template_file2)
    
    print(f"目标文件: {os.path.basename(result_path1)}")
    print("=" * 50)
    print(f"目标文件: {os.path.basename(result_path2)}")
    print("=" * 50)
	
    
    # 步骤3：执行数据复制
    try:
        print(f"\n📊 正在处理源文件: {os.path.basename(source_file1)}")
        
        copy_ranges(12, 3, source_file1, result_path1)
        copy_j_ranges(source_file1, result_path1)
        
        print("=" * 50)
        print("🎉 全部处理完成!")
        print(f"📁 结果文件: {result_path1}")
    except Exception as e:
        print(f"\n❌ 运行出错: {e}")

    try:
        print(f"\n📊 正在处理源文件: {os.path.basename(source_file2)}")
        
        copy_ranges(12, 7, source_file2, result_path2)
        adjust_file2_data(result_path2)
        
        print("=" * 50)
        print("🎉 全部处理完成!")
        print(f"📁 结果文件: {result_path2}")
    except Exception as e:
        print(f"\n❌ 运行出错: {e}")
        sys.exit(1)
if __name__ == "__main__":
    main()
